# -*- coding: utf-8 -*-
"""
RouteA v3: backfill BIG/SUB unit by (WAVE + META_ID + META_SEQ)
- 扫描 raw_root 下所有 xlsx：
    key = (WAVE, FILE_ID, SEQ) -> BIG_UNIT, SUB_UNIT, SRC_FILENAME
  FILE_ID 规则：
    1) 文件名开头的数字（>=6位）优先
    2) 否则用规范化后的文件名作为 FILE_ID（用于 24Q4 那种“处理后数据”文件）
- 从 DB table 读：
    PERSON_ID, WAVE, META_ID, META_FILE, META_SEQ, (可选既有单位列)
  生成 DB_FILE_ID：
    A) 若 META_ID 有数字（>=6位）→ FILE_ID = 该数字
    B) 否则从 META_FILE 文件名里提取开头数字；没有则用规范化文件名
    C) 再兜底：从 META_SOURCEDETAIL/META_SOURCE 抓取数字（若存在列）
- JOIN 后写：
    TABLE: unit_hierarchy_map_routeA_v3
    VIEW : <table>_routeA_v3  (新增 BIG_UNIT_A3/SUB_UNIT_A3 + BIG_UNIT_FINAL/SUB_UNIT_FINAL)

不改原表/原 view，旧脚本不受影响。
"""

import argparse, json, re, sqlite3
from pathlib import Path
import pandas as pd

BAD = {"", "nan", "none", "null", "#null!", "NULL", "None", "NaN"}

def norm_str(x):
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() in BAD else s

def norm_file(x):
    s = norm_str(x)
    if not s:
        return ""
    s = s.replace("\\", "/").split("/")[-1]
    s = re.sub(r"\s+", "", s)
    return s

def norm_seq(x):
    s = norm_str(x)
    if not s:
        return ""
    s = re.sub(r"[^\d]+", "", s)
    s = s.lstrip("0") or "0"
    return s

def extract_id_any(s: str) -> str:
    s = norm_str(s)
    if not s:
        return ""
    # 抓 >=6 位数字（问卷ID一般很长）
    m = re.search(r"(\d{6,})", s)
    return m.group(1) if m else ""

def extract_id_from_filename(fn: str) -> str:
    fn = norm_file(fn)
    if not fn:
        return ""
    # 文件名开头数字优先
    m = re.match(r"^(\d{6,})", fn)
    if m:
        return m.group(1)
    # 否则尝试任意位置数字
    return extract_id_any(fn)

def infer_wave_from_path(p: Path) -> str:
    s = str(p).replace("\\", "/")
    if "/24年1季度/" in s: return "24Q1"
    if "/24年2季度/" in s: return "24Q2"
    if "/24年3季度/" in s: return "24Q3"
    if "/24年4季度/" in s: return "24Q4"
    if "/25年1季度/" in s: return "25Q1"
    if "/25年2季度/" in s: return "25Q2"
    if "/25年3季度/" in s: return "25Q3"
    if "/25年4季度/" in s: return "25Q4"
    return ""

def bigunit_from_filename(fn: str) -> str:
    f = fn
    if "国家西南区域应急救援中心" in f: return "国家西南区域应急救援中心"
    if "重庆" in f and ("机动" in f or "机动队伍" in f or "机动部队" in f): return "国家消防救援局重庆机动队伍"
    if "阿坝" in f: return "四川省森林消防总队阿坝支队"
    if "攀枝花" in f: return "四川省森林消防总队攀枝花支队"
    if "甘孜" in f: return "四川省森林消防总队甘孜支队"
    if "凉山" in f: return "四川省森林消防总队凉山支队"
    if "机关" in f: return "四川省森林消防总队机关"
    if "省森林消防总队" in f and "支队" not in f: return "四川省森林消防总队"
    return ""

def pick_col(cols, keys):
    for k in keys:
        for c in cols:
            if k in str(c):
                return c
    return None

def read_raw_minimal(xlsx: Path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        header = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            header.append("" if cell is None else str(cell))
        cols = header
        col_seq  = pick_col(cols, ["序号"])
        col_post = pick_col(cols, ["您的工作岗位"])
        col_duty = pick_col(cols, ["您的职务"])
        if not col_seq:
            continue
        need = [c for c in [col_seq, col_post, col_duty] if c]
        df = pd.read_excel(xlsx, sheet_name=sh, usecols=need, engine="openpyxl", dtype=str)
        return df, {"sheet": sh, "col_seq": col_seq, "col_post": col_post, "col_duty": col_duty}
    return None, None

def pick_nonempty(s):
    s2 = s.fillna("").astype(str).map(norm_str)
    s2 = s2[s2 != ""]
    return s2.iloc[0] if len(s2) else ""

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--table", required=True)
    ap.add_argument("--raw_root", required=True)
    ap.add_argument("--out_dir", required=True)
    ap.add_argument("--map_table", default="unit_hierarchy_map_routeA_v3")
    ap.add_argument("--view_name", default="")
    args = ap.parse_args()

    out_dir = Path(args.out_dir); out_dir.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(args.db)

    cols = [r[1] for r in con.execute(f"PRAGMA table_info({args.table})").fetchall()]
    colset = set(cols)
    for need in ["PERSON_ID","WAVE","META_SEQ"]:
        if need not in colset:
            raise RuntimeError(f"[FATAL] {args.table} 缺少列: {need}")

    # 可选列（兜底用）
    has_meta_id = "META_ID" in colset
    has_meta_file = "META_FILE" in colset
    has_src = "META_SOURCE" in colset
    has_srcd = "META_SOURCEDETAIL" in colset

    # 找“已有单位列”，用于 FINAL
    big_exist = None
    for c in ["BIG_UNIT_FINAL","BIG_UNIT_A","BIG_UNIT_V5_3_2","BIG_UNIT_V5","BIG_UNIT_V4","BIG_UNIT_V2","BIG_UNIT_V1","BIG_UNIT"]:
        if c in colset:
            big_exist = c; break
    sub_exist = None
    for c in ["SUB_UNIT_FINAL","SUB_UNIT_A","SUB_UNIT_V5_3_2","SUB_UNIT_V5","SUB_UNIT_V4","SUB_UNIT_V2","SUB_UNIT_V1","SUB_UNIT","UNIT__FILLED","DEMO_UNITDEPT"]:
        if c in colset:
            sub_exist = c; break

    sel = ["PERSON_ID","WAVE","META_SEQ"]
    if has_meta_id: sel.append("META_ID")
    if has_meta_file: sel.append("META_FILE")
    if has_src: sel.append("META_SOURCE")
    if has_srcd: sel.append("META_SOURCEDETAIL")
    if big_exist: sel.append(big_exist)
    if sub_exist: sel.append(sub_exist)

    base = pd.read_sql_query(f"SELECT {', '.join(sel)} FROM {args.table}", con)
    base["WAVE"] = base["WAVE"].astype(str).map(norm_str)
    base["SEQ_N"] = base["META_SEQ"].map(norm_seq)

    # DB_FILE_ID：优先 META_ID（抓数字），再 META_FILE（抓文件名数字/规范名），再 source/detail
    def build_db_file_id(row):
        if has_meta_id:
            mid = extract_id_any(row.get("META_ID"))
            if mid: return mid
        if has_meta_file:
            mf = norm_file(row.get("META_FILE"))
            if mf:
                fid = extract_id_from_filename(mf)
                return fid if fid else mf  # 无数字则用文件名
        # 兜底：从 source/detail 抓数字
        if has_srcd:
            x = extract_id_any(row.get("META_SOURCEDETAIL"))
            if x: return x
        if has_src:
            x = extract_id_any(row.get("META_SOURCE"))
            if x: return x
        return ""

    base["FILE_ID_DB"] = base.apply(build_db_file_id, axis=1)

    # 扫 raw -> raw_map(key=(WAVE, FILE_ID, SEQ))
    raw_root = Path(args.raw_root)
    xls = [p for p in raw_root.rglob("*.xlsx") if "~$" not in p.name]
    rows = []
    meta = []
    for fp in xls:
        wave = infer_wave_from_path(fp)
        if not wave:
            continue
        df, info = read_raw_minimal(fp)
        if df is None or df.empty:
            continue
        src = fp.name
        src_n = norm_file(src)
        fid = extract_id_from_filename(src_n)
        file_id = fid if fid else src_n  # 没数字就用文件名
        big = bigunit_from_filename(src_n)

        seq_col = info["col_seq"]
        post_col = info["col_post"]
        duty_col = info["col_duty"]

        for _, r in df.iterrows():
            seq = norm_seq(r.get(seq_col))
            if not seq:
                continue
            sub = norm_str(r.get(post_col)) if post_col else ""
            if not sub and duty_col:
                sub = norm_str(r.get(duty_col))
            rows.append({
                "WAVE": wave,
                "FILE_ID_RAW": file_id,
                "SEQ": seq,
                "BIG_UNIT_A3": big,
                "SUB_UNIT_A3": sub,
                "SRC_FILENAME_A3": src,
            })
        meta.append({"file": str(fp), "wave": wave, "file_id_raw": file_id, **info, "rows": int(len(df))})

    raw_map = pd.DataFrame(rows)
    if raw_map.empty:
        raise RuntimeError("[FATAL] raw_root 扫描不到包含“序号”的 sheet，无法回挂。")

    raw_map2 = raw_map.groupby(["WAVE","FILE_ID_RAW","SEQ"], as_index=False).agg(
        BIG_UNIT_A3=("BIG_UNIT_A3", pick_nonempty),
        SUB_UNIT_A3=("SUB_UNIT_A3", pick_nonempty),
        SRC_FILENAME_A3=("SRC_FILENAME_A3", pick_nonempty),
    )

    # join：WAVE + FILE_ID + SEQ
    joined = base.merge(
        raw_map2,
        left_on=["WAVE","FILE_ID_DB","SEQ_N"],
        right_on=["WAVE","FILE_ID_RAW","SEQ"],
        how="left"
    )

    hit = joined["BIG_UNIT_A3"].fillna("").map(norm_str) != ""
    out = joined.loc[hit, ["PERSON_ID","WAVE","FILE_ID_DB","SRC_FILENAME_A3","BIG_UNIT_A3","SUB_UNIT_A3"]].copy()
    out["MATCH_METHOD_A3"] = "META_ID_SEQ"
    out = out.drop_duplicates(subset=["PERSON_ID","WAVE"], keep="first")

    # 写 mapping 表
    con.execute(f"DROP TABLE IF EXISTS {args.map_table}")
    out.to_sql(args.map_table, con, if_exists="replace", index=False)
    con.execute(f"CREATE INDEX IF NOT EXISTS idx_{args.map_table}_pw ON {args.map_table}(PERSON_ID, WAVE)")

    # 新 VIEW
    view_name = args.view_name.strip() or f"{args.table}_routeA_v3"
    con.execute(f"DROP VIEW IF EXISTS {view_name}")

    if big_exist:
        big_final = f"COALESCE(NULLIF(TRIM(a.{big_exist}),''), NULLIF(TRIM(m.BIG_UNIT_A3),''))"
    else:
        big_final = f"NULLIF(TRIM(m.BIG_UNIT_A3),'')"

    if sub_exist:
        sub_final = f"COALESCE(NULLIF(TRIM(a.{sub_exist}),''), NULLIF(TRIM(m.SUB_UNIT_A3),''))"
    else:
        sub_final = f"NULLIF(TRIM(m.SUB_UNIT_A3),'')"

    con.execute(f"""
        CREATE VIEW {view_name} AS
        SELECT a.*,
               m.FILE_ID_DB,
               m.SRC_FILENAME_A3,
               m.BIG_UNIT_A3,
               m.SUB_UNIT_A3,
               m.MATCH_METHOD_A3,
               {big_final} AS BIG_UNIT_FINAL_A3,
               {sub_final} AS SUB_UNIT_FINAL_A3
        FROM {args.table} a
        LEFT JOIN {args.map_table} m
          ON a.PERSON_ID = m.PERSON_ID AND a.WAVE = m.WAVE
    """)
    con.commit()
    con.close()

    # 报告
    summary = {
        "db": args.db,
        "table": args.table,
        "view_created": view_name,
        "map_table": args.map_table,
        "raw_files_scanned": len(meta),
        "filled_rows": int(len(out)),
        "filled_rate_all": float(len(out) / max(1, len(base))),
        "note": "v3 用 META_ID(或从 META_FILE/source/detail 提取数字) + META_SEQ 回挂 raw。"
    }
    (out_dir / "SUMMARY_ROUTEA_V3.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    pd.DataFrame(meta).to_csv(out_dir / "raw_read_meta_routeA_v3.csv", index=False, encoding="utf-8-sig")

    print("================================================================================")
    print("[OK] RouteA v3 done (by WAVE + META_ID + META_SEQ).")
    print(f"[OK] VIEW: {view_name}")
    print(f"[OK] filled_rows={len(out)}  filled_rate_all={len(out)/max(1,len(base)):.4f}")
    print(f"[OK] reports: {out_dir}")
    print("================================================================================")

if __name__ == "__main__":
    main()
