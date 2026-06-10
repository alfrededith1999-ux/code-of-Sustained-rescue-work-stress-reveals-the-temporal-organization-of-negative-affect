# -*- coding: utf-8 -*-
import argparse, re, json, sqlite3
from pathlib import Path
import pandas as pd

BAD = {"", "nan", "none", "null", "#null!", "NULL", "None", "NaN"}

def norm_str(x):
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() in BAD else s

def canon_phone(x):
    s = norm_str(x)
    if not s:
        return ""
    digits = re.sub(r"\D+", "", s)
    return digits[-11:] if len(digits) >= 11 else digits

def canon_name(x):
    s = norm_str(x)
    if not s:
        return ""
    return re.sub(r"\s+", "", s)

def norm_file(x):
    s = norm_str(x)
    if not s:
        return ""
    s = s.replace("\\", "/")
    s = s.split("/")[-1]
    s = re.sub(r"\s+", "", s)
    return s

def norm_seq(x):
    s = norm_str(x)
    if not s:
        return ""
    # 可能是 1.0 / '001' / ' 1 '
    s = re.sub(r"[^\d]+", "", s)
    s = s.lstrip("0") or "0"
    return s

def infer_wave_from_path(p: Path) -> str:
    s = str(p).replace("\\", "/")
    if "/24年1季度/" in s: return "24Q1"
    if "/24年2季度/" in s: return "24Q2"
    if "/24年3季度/" in s: return "24Q3"
    if "/24年4季度/" in s: return "24Q4"
    if "/25年1季度/" in s: return "25Q1"
    if "/25年2季度/" in s: return "25Q2"
    if "/25年3季度/" in s: return "25Q3"
    if "/25年4季度/" in s: return "25Q4"
    return ""

def bigunit_from_filename(fn: str) -> str:
    f = fn
    # 你给的“人工名单逻辑”的关键词归并（可按你实际口径再改）
    if "国家西南区域应急救援中心" in f: return "国家西南区域应急救援中心"
    if "重庆" in f and ("机动" in f or "机动队伍" in f or "机动部队" in f): return "国家消防救援局重庆机动队伍"
    if "阿坝" in f: return "四川省森林消防总队阿坝支队"
    if "攀枝花" in f: return "四川省森林消防总队攀枝花支队"
    if "甘孜" in f: return "四川省森林消防总队甘孜支队"
    if "凉山" in f: return "四川省森林消防总队凉山支队"
    if "机关" in f: return "四川省森林消防总队机关"
    if "省森林消防总队" in f and "支队" not in f: return "四川省森林消防总队"
    return ""

def pick_col(cols, contains_any):
    for key in contains_any:
        for c in cols:
            if key in str(c):
                return c
    return None

def read_raw_minimal(xlsx: Path):
    # 尽量稳：先读表头，找含关键字的列，再按存在列读取
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        header = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            header.append("" if cell is None else str(cell))
        cols = header

        col_seq   = pick_col(cols, ["序号"])
        col_name  = pick_col(cols, ["您的姓名"])
        col_phone = pick_col(cols, ["您的联系电话"])
        # sub unit 常常在 “工作岗位/职务” 里体现（你之前 UNIT__FILLED 就是这路子）
        col_post  = pick_col(cols, ["您的工作岗位"])
        col_duty  = pick_col(cols, ["您的职务"])

        need = [c for c in [col_seq, col_name, col_phone, col_post, col_duty] if c]
        if not col_seq or not need:
            continue

        df = pd.read_excel(xlsx, sheet_name=sh, usecols=need, engine="openpyxl", dtype=str)
        return df, {"sheet": sh, "col_seq": col_seq, "col_name": col_name, "col_phone": col_phone, "col_post": col_post, "col_duty": col_duty}

    return None, None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--table", required=True)  # 建议: assessment_wide_trackable_dedup_unitfilled
    ap.add_argument("--raw_root", required=True)
    ap.add_argument("--out_dir", required=True)
    ap.add_argument("--map_table", default="unit_hierarchy_map_routeA_v2")
    ap.add_argument("--view_name", default="")
    args = ap.parse_args()

    out_dir = Path(args.out_dir); out_dir.mkdir(parents=True, exist_ok=True)

    con = sqlite3.connect(args.db)

    # 找 base 表里的 meta_file / meta_seq 列名
    cols = [r[1] for r in con.execute(f"PRAGMA table_info({args.table})").fetchall()]
    colset = set(cols)

    required = ["PERSON_ID", "WAVE"]
    for r in required:
        if r not in colset:
            raise RuntimeError(f"[FATAL] base table 缺少必需列: {r}")

    meta_file_col = None
    for c in ["META_FILE", "meta_file", "meta_file_name", "meta_file_basename"]:
        if c in colset:
            meta_file_col = c; break

    meta_seq_col = None
    for c in ["META_SEQ", "meta_seq", "meta_seq_id", "meta_seqno"]:
        if c in colset:
            meta_seq_col = c; break

    if not meta_file_col or not meta_seq_col:
        raise RuntimeError(f"[FATAL] base table 找不到 meta_file/meta_seq。检测到 meta_file={meta_file_col}, meta_seq={meta_seq_col}. 你需要先确认表里是否有 META_FILE/META_SEQ。")

    # 只取缺失 BIG_UNIT 的行（你现在 BIG_UNIT 叫 BIG_UNIT_V5_3_2 或 BIG_UNIT_V5_3_2_routeA_v1 里 BIG_UNIT_A）
    # 这里直接检查 UNIT__FILLED 与 DEMO_PHONE_CANON 为空的那些行，优先补“键”
    phone_col = "DEMO_PHONE_CANON" if "DEMO_PHONE_CANON" in colset else None
    unit_col  = "UNIT__FILLED" if "UNIT__FILLED" in colset else None

    sel = ["PERSON_ID", "WAVE", meta_file_col, meta_seq_col]
    if phone_col: sel.append(phone_col)
    if unit_col:  sel.append(unit_col)

    base = pd.read_sql_query(f"SELECT {', '.join(sel)} FROM {args.table}", con)
    base["WAVE"] = base["WAVE"].astype(str).map(norm_str)
    base["META_FILE_N"] = base[meta_file_col].map(norm_file)
    base["META_SEQ_N"]  = base[meta_seq_col].map(norm_seq)

    if phone_col:
        base["PHONE_CANON"] = base[phone_col].map(canon_phone)
    else:
        base["PHONE_CANON"] = ""

    if unit_col:
        base["UNIT_FILLED"] = base[unit_col].map(norm_str)
    else:
        base["UNIT_FILLED"] = ""

    need_mask = (base["PHONE_CANON"]=="") & (base["UNIT_FILLED"]=="")
    base_need = base[need_mask & (base["WAVE"]!="") & (base["META_FILE_N"]!="") & (base["META_SEQ_N"]!="")].copy()

    # 扫 raw 建 (WAVE, SRC_FILE, SEQ) -> phone/name/sub/big
    raw_root = Path(args.raw_root)
    xls = [p for p in raw_root.rglob("*.xlsx") if "~$" not in p.name]
    rows = []
    meta = []

    for fp in xls:
        wave = infer_wave_from_path(fp)
        if not wave:
            continue
        df, info = read_raw_minimal(fp)
        if df is None or df.empty:
            continue
        src = fp.name
        big = bigunit_from_filename(src)

        # 提取
        seq_col = info["col_seq"]
        name_col = info["col_name"]
        phone_col_r = info["col_phone"]
        post_col = info["col_post"]
        duty_col = info["col_duty"]

        for _, r in df.iterrows():
            seq = norm_seq(r.get(seq_col))
            if not seq:
                continue
            name = canon_name(r.get(name_col)) if name_col else ""
            phone = canon_phone(r.get(phone_col_r)) if phone_col_r else ""
            # SUB 单位：优先岗位，再职务（按你之前逻辑）
            sub = norm_str(r.get(post_col)) if post_col else ""
            if not sub and duty_col:
                sub = norm_str(r.get(duty_col))
            rows.append({
                "WAVE": wave,
                "SRC_FILE": src,
                "SRC_FILE_N": norm_file(src),
                "SEQ": seq,
                "RAW_PHONE": phone,
                "RAW_NAME": name,
                "RAW_SUB": sub,
                "RAW_BIG": big
            })
        meta.append({"file": str(fp), "wave": wave, **info, "rows": int(len(df))})

    raw_map = pd.DataFrame(rows)
    if raw_map.empty:
        raise RuntimeError("[FATAL] raw_root 扫描不到可用 raw 映射（没读到包含“序号”的sheet）。")

    # raw_map 去重：对 (WAVE, SRC_FILE_N, SEQ) 取“非空优先”
    def pick_nonempty(s):
        s2 = s.fillna("").astype(str).map(norm_str)
        s2 = s2[s2 != ""]
        return s2.iloc[0] if len(s2) else ""

    raw_map = raw_map.sort_values(["WAVE","SRC_FILE_N","SEQ"])
    raw_map2 = raw_map.groupby(["WAVE","SRC_FILE_N","SEQ"], as_index=False).agg(
        RAW_PHONE=("RAW_PHONE", pick_nonempty),
        RAW_NAME=("RAW_NAME", pick_nonempty),
        RAW_SUB=("RAW_SUB", pick_nonempty),
        RAW_BIG=("RAW_BIG", pick_nonempty)
    )

    # 回挂到 base_need：按 (WAVE, META_FILE_N, META_SEQ_N) join raw_map2
    joined = base_need.merge(
        raw_map2,
        left_on=["WAVE","META_FILE_N","META_SEQ_N"],
        right_on=["WAVE","SRC_FILE_N","SEQ"],
        how="left"
    )

    # 生成 mapping 表（只保留命中的）
    hit = joined["RAW_BIG"].fillna("").map(norm_str) != ""
    out = joined.loc[hit, ["PERSON_ID","WAVE","META_FILE_N","META_SEQ_N","RAW_PHONE","RAW_NAME","RAW_SUB","RAW_BIG"]].copy()
    out = out.rename(columns={
        "META_FILE_N":"SRC_FILENAME_A2",
        "RAW_PHONE":"PHONE_CANON_A2",
        "RAW_NAME":"NAME_CANON_A2",
        "RAW_SUB":"SUB_UNIT_A2",
        "RAW_BIG":"BIG_UNIT_A2"
    })
    out["MATCH_METHOD_A2"] = "META_FILE_SEQ"

    # 写表
    con.execute(f"DROP TABLE IF EXISTS {args.map_table}")
    out.to_sql(args.map_table, con, if_exists="replace", index=False)
    con.execute(f"CREATE INDEX IF NOT EXISTS idx_{args.map_table}_pw ON {args.map_table}(PERSON_ID, WAVE)")

    # 新 VIEW：不改原字段，新增 *_A2 字段；BIG_UNIT 最终建议你用 COALESCE(旧, A2)
    view_name = args.view_name.strip() or f"{args.table}_routeA_v2"
    con.execute(f"DROP VIEW IF EXISTS {view_name}")
    con.execute(f"""
        CREATE VIEW {view_name} AS
        SELECT a.*,
               m.SRC_FILENAME_A2,
               m.PHONE_CANON_A2,
               m.NAME_CANON_A2,
               m.SUB_UNIT_A2,
               m.BIG_UNIT_A2,
               m.MATCH_METHOD_A2
        FROM {args.table} a
        LEFT JOIN {args.map_table} m
          ON a.PERSON_ID = m.PERSON_ID AND a.WAVE = m.WAVE
    """)
    con.commit()
    con.close()

    # 报告
    summary = {
        "db": args.db,
        "base_table": args.table,
        "map_table": args.map_table,
        "view_created": view_name,
        "raw_files_scanned": len(meta),
        "base_need_rows": int(len(base_need)),
        "filled_rows": int(len(out)),
        "filled_rate_on_need": float(len(out) / max(1, len(base_need))),
        "note": "这一步只负责把 raw 的 BIG/SUB/PHONE/NAME 回挂成 *_A2 字段，不覆盖原字段；后续你用 COALESCE 做最终单位字段。"
    }
    (out_dir / "SUMMARY_ROUTEA_V2.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    pd.DataFrame(meta).to_csv(out_dir / "raw_read_meta_routeA_v2.csv", index=False, encoding="utf-8-sig")

    print("================================================================================")
    print("[OK] RouteA v2 backfill done (by WAVE + meta_file + meta_seq).")
    print(f"[OK] VIEW: {view_name}")
    print(f"[OK] base_need_rows={len(base_need)}  filled_rows={len(out)}  filled_rate_on_need={len(out)/max(1,len(base_need)):.4f}")
    print(f"[OK] reports: {out_dir}")
    print("================================================================================")

if __name__ == "__main__":
    main()
